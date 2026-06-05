# apps/imports/management/commands/import_risposte_eg.py
"""Import diari da Excel Jotform 'EG - Diario della Specialità di Squadriglia'.

Fogli:
  Risposte EG  → Anagrafica, Presentazione, MembroSq, Impresa 1/2, Missione,
                 PostoAzione, EsitoSpecialita; stato → RELAZIONE_FINALE
  Risposte staff → RelazioneFinale; stato → INVIATO

Opzione --importa-foto:
  Legge le sottocartelle Drive (una per diario, formato
  'Zona X_Gruppo_Reparto_NomeSquadriglia'), abbina i file ai moduli
  usando i filename presenti negli URL dell'Excel, crea record Allegato.

Uso:
  uv run python manage.py import_risposte_eg /percorso/file.xlsx
  uv run python manage.py import_risposte_eg /percorso/file.xlsx --dry-run
  uv run python manage.py import_risposte_eg /percorso/file.xlsx --solo-staff
  uv run python manage.py import_risposte_eg /percorso/file.xlsx --edizione 1
  uv run python manage.py import_risposte_eg /percorso/file.xlsx \\
    --importa-foto --foto-folder-id <DRIVE_FOLDER_ID> --foto-account <EMAIL>
"""
from __future__ import annotations

import re
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

# ---------------------------------------------------------------------------
# Mappature
# ---------------------------------------------------------------------------

_RUOLO_MAP = {
    "capo sq": "csq",
    "vice capo sq": "vcsq",
    "squadrigliere": "squadrigliere",
}

_SENTIERO_MAP = {
    "responsabilità": "responsabilita",
    "responsabilita": "responsabilita",
    "competenza": "competenza",
    "scoperta": "scoperta",
}

_STATO_SPECIALITA_MAP = {
    "conquistata": "conquistata",
    "incammino": "in_cammino",
    "in cammino": "in_cammino",
    "non conquistata": "non_conquistata",
}


# ---------------------------------------------------------------------------
# Helpers parsing
# ---------------------------------------------------------------------------

def _val(row: tuple, idx: int) -> str:
    """Restituisce il valore della colonna (1-based) come stringa pulita."""
    v = row[idx - 1]
    return str(v).strip() if v is not None else ""


def _date(row: tuple, idx: int) -> date | None:
    v = row[idx - 1]
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_blocchi(testo: str) -> list[dict]:
    """Parsa righe del tipo 'Chiave: valore, Chiave2: valore2' separate da \\n."""
    risultati = []
    if not testo:
        return risultati
    for riga in testo.split("\n"):
        riga = riga.strip()
        if not riga:
            continue
        d: dict[str, str] = {}
        # Trova tutte le coppie chiave: valore
        for m in re.finditer(r"([A-Za-zÀ-ÿ' ]+?):\s*([^,]+?)(?=,\s*[A-Za-zÀ-ÿ' ]+?:|$)", riga):
            d[m.group(1).strip().lower()] = m.group(2).strip()
        if d:
            risultati.append(d)
    return risultati


def _primo_url(testo: str) -> str:
    """Estrae il primo URL da una stringa che può contenere URL multipli."""
    if not testo:
        return ""
    urls = [u.strip() for u in testo.split("\n") if u.strip().startswith("http")]
    return urls[0] if urls else ""


# ---------------------------------------------------------------------------
# Ricerca Diario
# ---------------------------------------------------------------------------

def _parse_drive_folder_name(nome: str) -> tuple[str, str, str, str] | None:
    """Parsa 'Zona X_Gruppo_Reparto_Squadriglia' → (zona_core, gruppo, reparto, sq).

    Split su '_' con maxsplit=3: zona (prima parte), gruppo, reparto, squadriglia.
    """
    parts = nome.split("_", 3)
    if len(parts) != 4:
        return None
    zona_raw, gruppo, reparto, squadriglia = parts
    zona_core = re.sub(r"^zona\s+", "", zona_raw.strip(), flags=re.IGNORECASE).strip()
    return zona_core, gruppo.strip(), reparto.strip(), squadriglia.strip()


def _normalizza_zona(zona: str) -> str:
    """Rimuove il prefisso 'Zona'/'ZONA' per il confronto."""
    return re.sub(r"^zona\s+", "", zona.strip(), flags=re.IGNORECASE).strip()


def _trova_diario(nome_sq: str, reparto: str, gruppo: str, zona: str, edizione):
    """Cerca il Diario con matching case-insensitive.

    Il nome della zona può avere prefisso 'Zona'/'ZONA' sia nell'Excel che nel DB;
    il matching usa icontains sulla parte significativa del nome.
    """

    from apps.diaries.models import Diario

    zona_core = _normalizza_zona(zona)

    qs = Diario.objects.filter(
        edizione=edizione,
        squadriglia__nome__iexact=nome_sq,
        squadriglia__reparto__nome__iexact=reparto,
        squadriglia__reparto__gruppo__nome__iexact=gruppo,
    ).filter(
        # icontains gestisce sia "Salerno" che "ZONA SALERNO"
        squadriglia__reparto__gruppo__zona__nome__icontains=zona_core,
    ).select_related(
        "squadriglia__reparto__gruppo__zona",
        "csq", "crp",
    )

    if qs.count() > 1:
        # Preferisce il match con nome esatto (es. evita "Zona Salerno Nord" per "Salerno")
        exact = qs.filter(squadriglia__reparto__gruppo__zona__nome__iexact=zona_core)
        if exact.exists():
            return exact.first()

    return qs.first()


# ---------------------------------------------------------------------------
# Import Risposte EG
# ---------------------------------------------------------------------------

def _import_riga_eg(row: tuple, edizione, dry_run: bool, verbosity: int) -> str:
    """Processa una riga del foglio Risposte EG. Restituisce 'ok'/'skip'/'err:msg'."""
    from apps.diaries.models import (
        Anagrafica,
        MembroSq,
        Missione,
        PostoAzioneMissione,
        Presentazione,
        StatoDiario,
        TipoDiario,
    )

    nome_sq  = _val(row, 2)
    reparto  = _val(row, 4)
    gruppo   = _val(row, 5)
    zona     = _val(row, 6)

    diario = _trova_diario(nome_sq, reparto, gruppo, zona, edizione)
    if not diario:
        return f"err:diario non trovato ({zona} / {gruppo} / {reparto} / {nome_sq})"

    if dry_run:
        return f"ok (dry-run): {diario}"

    with transaction.atomic():
        sp = transaction.savepoint()
        try:
            # --- Anagrafica ---
            ana, _ = Anagrafica.objects.get_or_create(diario=diario)
            ana.specialita    = _val(row, 7)
            ana.crp_nome      = _val(row, 13)
            ana.crp_cognome   = _val(row, 14)
            ana.crp_email     = _val(row, 15)
            ana.crp_cell      = _val(row, 16)
            ana.save()

            # Tipo diario
            tipo_raw = _val(row, 8).lower()
            if tipo_raw == "rinnovo":
                diario.tipo = TipoDiario.RINNOVO
            else:
                diario.tipo = TipoDiario.NUOVO
            diario.save(update_fields=["tipo"])

            # --- Presentazione ---
            pres, _ = Presentazione.objects.get_or_create(diario=diario)
            pres.cosa_sappiamo_fare = _val(row, 18)
            pres.save()

            # --- Membri ---
            MembroSq.objects.filter(presentazione=pres).delete()
            for b in _parse_blocchi(_val(row, 17)):
                ruolo_raw = b.get("ruolo", "").lower()
                sentiero_raw = b.get("sentiero", "").lower()
                nome_completo = b.get("nome", "").strip()
                parti = nome_completo.split(" ", 1)
                MembroSq.objects.create(
                    presentazione=pres,
                    nome=parti[0] if parti else nome_completo,
                    cognome=parti[1] if len(parti) > 1 else "",
                    ruolo=_RUOLO_MAP.get(ruolo_raw, "altro"),
                    sentiero=_SENTIERO_MAP.get(sentiero_raw, "non_specificato"),
                )

            # --- Impresa 1 ---
            _import_impresa(diario, numero=1, row=row,
                            col_titolo=19, col_inizio=20, col_fine=21,
                            col_perche=22, col_come=23, col_cosa=24,
                            col_posti=25, col_specialita=27, col_brevetti=28,
                            col_foto=29, col_video=30)

            # --- Impresa 2 (solo se c'è titolo) ---
            if _val(row, 31):
                _import_impresa(diario, numero=2, row=row,
                                col_titolo=31, col_inizio=32, col_fine=33,
                                col_perche=34, col_come=35, col_cosa=36,
                                col_posti=37, col_specialita=39, col_brevetti=40,
                                col_foto=41, col_video=42)

            # --- Missione (solo se c'è titolo) ---
            if _val(row, 43):
                miss, _ = Missione.objects.get_or_create(diario=diario)
                miss.titolo                 = _val(row, 43)
                miss.data                   = _date(row, 44)
                miss.descrizione_svolgimento = _val(row, 45)
                miss.save()
                PostoAzioneMissione.objects.filter(missione=miss).delete()
                # La missione non ha una colonna "Posti d'azione" separata nell'Excel

            # --- Transizione stato ---
            if diario.stato == StatoDiario.IN_COMPILAZIONE:
                diario.csq_invia()

            transaction.savepoint_commit(sp)
            return f"ok: {diario}"

        except Exception as exc:
            transaction.savepoint_rollback(sp)
            return f"err:{exc}"


def _import_impresa(diario, numero: int, row: tuple,
                    col_titolo, col_inizio, col_fine,
                    col_perche, col_come, col_cosa,
                    col_posti, col_specialita, col_brevetti,
                    col_foto, col_video):
    from apps.diaries.models import (
        EsitoSpecialita,
        Impresa,
        PostoAzione,
        TipoEsito,
    )

    imp, _ = Impresa.objects.get_or_create(diario=diario, numero=numero)
    imp.titolo      = _val(row, col_titolo)
    imp.data_inizio = _date(row, col_inizio)
    imp.data_fine   = _date(row, col_fine)
    imp.perche      = _val(row, col_perche)
    imp.come        = _val(row, col_come)
    imp.cosa        = _val(row, col_cosa)
    url_video = _primo_url(_val(row, col_video))
    if not url_video:
        url_video = _primo_url(_val(row, col_foto))
    imp.link_esterno = url_video
    imp.save()

    # Posti d'azione
    PostoAzione.objects.filter(impresa=imp).delete()
    for b in _parse_blocchi(_val(row, col_posti)):
        posto = b.get("posto d'azione", b.get("posto", "")).strip()
        nome  = b.get("nome", "").strip()
        if posto:
            desc = f"{nome} — {posto}" if nome else posto
            PostoAzione.objects.create(impresa=imp, descrizione=desc[:300])

    # Specialità individuali
    EsitoSpecialita.objects.filter(impresa=imp, tipo=TipoEsito.SPECIALITA).delete()
    for b in _parse_blocchi(_val(row, col_specialita)):
        nome  = b.get("specialità", b.get("nome", "")).strip()
        stato_raw = b.get("conquistata", "").strip().lower()
        stato = _STATO_SPECIALITA_MAP.get(stato_raw, "in_cammino")
        if nome:
            EsitoSpecialita.objects.create(
                impresa=imp, tipo=TipoEsito.SPECIALITA,
                nome=nome, stato=stato,
            )

    # Brevetti
    EsitoSpecialita.objects.filter(impresa=imp, tipo=TipoEsito.BREVETTO).delete()
    for b in _parse_blocchi(_val(row, col_brevetti)):
        nome  = b.get("brevetto", b.get("nome", "")).strip()
        stato_raw = b.get("conquistata", "").strip().lower()
        stato = _STATO_SPECIALITA_MAP.get(stato_raw, "in_cammino")
        if nome:
            EsitoSpecialita.objects.create(
                impresa=imp, tipo=TipoEsito.BREVETTO,
                nome=nome, stato=stato,
            )


# ---------------------------------------------------------------------------
# Import Risposte staff
# ---------------------------------------------------------------------------

def _import_riga_staff(row: tuple, edizione, dry_run: bool, verbosity: int) -> str:
    from apps.diaries.models import RelazioneFinale, StatoDiario

    nome_sq = _val(row, 5)
    reparto = _val(row, 7)
    gruppo  = _val(row, 8)
    zona    = _val(row, 9)

    diario = _trova_diario(nome_sq, reparto, gruppo, zona, edizione)
    if not diario:
        return f"err:diario non trovato ({zona} / {gruppo} / {reparto} / {nome_sq})"

    if dry_run:
        return f"ok (dry-run): {diario}"

    with transaction.atomic():
        sp = transaction.savepoint()
        try:
            rel, _ = RelazioneFinale.objects.get_or_create(diario=diario)
            rel.sintesi_impresa_1  = _val(row, 11)
            rel.sintesi_impresa_2  = _val(row, 12)
            rel.sintesi_missione   = _val(row, 13)
            rel.considerazioni     = _val(row, 14)
            conquistata_raw = _val(row, 15).strip().lower()
            if conquistata_raw == "si":
                rel.specialita_conquistata = True
            elif conquistata_raw == "no":
                rel.specialita_conquistata = False
            rel.save()

            if diario.stato == StatoDiario.RELAZIONE_FINALE:
                diario.invia()

            transaction.savepoint_commit(sp)
            return f"ok: {diario}"

        except Exception as exc:
            transaction.savepoint_rollback(sp)
            return f"err:{exc}"


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa diari da Excel Jotform 'EG - Diario della Specialità di Squadriglia'."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Percorso del file Excel (.xlsx)")
        parser.add_argument(
            "--edizione", type=int, default=None,
            help="PK dell'edizione target (default: edizione aperta più recente).",
        )
        parser.add_argument(
            "--solo-staff", action="store_true",
            help="Importa solo il foglio 'Risposte staff' (relazioni finali).",
        )
        parser.add_argument(
            "--solo-eg", action="store_true",
            help="Importa solo il foglio 'Risposte EG' (diari squadriglia).",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Stampa le operazioni senza salvare nulla.",
        )
        parser.add_argument(
            "--importa-foto", action="store_true",
            help="Importa le foto da Google Drive (richiede --foto-folder-id e --foto-account).",
        )
        parser.add_argument(
            "--foto-folder-id",
            help="ID cartella Drive padre contenente le sottocartelle per diario.",
        )
        parser.add_argument(
            "--foto-account",
            help="Email account Google Drive (deve essere in DriveCredenziali).",
        )

    def handle(self, *args, **options):
        import openpyxl

        from apps.editions.models import Edizione, StatoEdizione

        file_path = options["file"]
        dry_run   = options["dry_run"]
        verbosity = options["verbosity"]

        # Edizione
        if options["edizione"]:
            try:
                edizione = Edizione.objects.get(pk=options["edizione"])
            except Edizione.DoesNotExist:
                raise CommandError(f"Edizione {options['edizione']} non trovata.") from None
        else:
            edizione = (
                Edizione.objects
                .filter(stato__in=[StatoEdizione.APERTA, StatoEdizione.IN_VALUTAZIONE])
                .order_by("-anno")
                .first()
            )
            if not edizione:
                raise CommandError("Nessuna edizione aperta trovata. Usa --edizione <pk>.")

        self.stdout.write(f"Edizione: {edizione}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessun dato verrà salvato."))

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File non trovato: {file_path}") from None
        except Exception as exc:
            raise CommandError(f"Errore apertura file: {exc}") from exc

        solo_staff = options["solo_staff"]
        solo_eg    = options["solo_eg"]

        # --- Foglio Risposte EG ---
        if not solo_staff:
            ws = wb["Risposte EG"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            ok = err = skip = 0
            self.stdout.write(f"\n→ Risposte EG ({len(rows)} righe)…")
            for i, row in enumerate(rows, 2):
                if not row[1]:  # Squadriglia vuota → riga vuota
                    continue
                risultato = _import_riga_eg(row, edizione, dry_run, verbosity)
                if risultato.startswith("ok"):
                    ok += 1
                    if verbosity >= 2:
                        self.stdout.write(f"  riga {i}: {risultato}")
                elif risultato.startswith("skip"):
                    skip += 1
                    if verbosity >= 2:
                        self.stdout.write(f"  riga {i}: {risultato}")
                else:
                    err += 1
                    self.stdout.write(self.style.ERROR(f"  riga {i}: {risultato}"))
            self.stdout.write(
                self.style.SUCCESS(f"  Risposte EG: {ok} importate, {skip} saltate, {err} errori.")
            )

        # --- Foglio Risposte staff ---
        if not solo_eg:
            ws2 = wb["Risposte staff"]
            rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
            ok2 = err2 = skip2 = 0
            self.stdout.write(f"\n→ Risposte staff ({len(rows2)} righe)…")
            for i, row in enumerate(rows2, 2):
                if not row[4]:  # Squadriglia vuota
                    continue
                risultato = _import_riga_staff(row, edizione, dry_run, verbosity)
                if risultato.startswith("ok"):
                    ok2 += 1
                    if verbosity >= 2:
                        self.stdout.write(f"  riga {i}: {risultato}")
                elif risultato.startswith("skip"):
                    skip2 += 1
                else:
                    err2 += 1
                    self.stdout.write(self.style.ERROR(f"  riga {i}: {risultato}"))
            self.stdout.write(
                self.style.SUCCESS(f"  Risposte staff: {ok2} importate, {skip2} saltate, {err2} errori.")
            )

        # --- Foto da Google Drive ---
        if options["importa_foto"]:
            if not options["foto_folder_id"] or not options["foto_account"]:
                raise CommandError(
                    "--importa-foto richiede --foto-folder-id e --foto-account."
                )
            self._importa_foto(
                wb=wb,
                edizione=edizione,
                folder_id=options["foto_folder_id"],
                account_email=options["foto_account"],
                dry_run=dry_run,
                verbosity=verbosity,
            )

        self.stdout.write(self.style.SUCCESS("\nImport completato."))

    # ------------------------------------------------------------------
    # Import foto da Google Drive
    # ------------------------------------------------------------------

    def _importa_foto(self, wb, edizione, folder_id, account_email,
                      dry_run, verbosity):
        from apps.diaries.models import Allegato, StatoSync
        from apps.storage_drive.models import DriveCredenziali
        from apps.storage_drive.service import _build_drive_service

        try:
            cred = DriveCredenziali.objects.get(account_email=account_email)
        except DriveCredenziali.DoesNotExist:
            raise CommandError(
                f"Account Drive '{account_email}' non trovato in DriveCredenziali. "
                "Connetti prima l'account tramite /impostazioni/ o /edizioni/<pk>/modifica/."
            ) from None

        service = _build_drive_service(cred)

        # 1. Costruisce la mappa filename → modulo dall'Excel
        modulo_map = self._build_filename_modulo_map(wb)
        self.stdout.write(
            f"\n→ Foto Drive (cartella: {folder_id}, account: {account_email})…"
        )

        # 2. Lista tutte le sottocartelle nel folder padre
        sottocartelle = self._lista_sottocartelle(service, folder_id)
        self.stdout.write(f"  {len(sottocartelle)} sottocartelle trovate.")

        ok = err = 0
        for cartella in sottocartelle:
            nome_cartella = cartella["name"]
            parsed = _parse_drive_folder_name(nome_cartella)
            if not parsed:
                self.stdout.write(
                    self.style.WARNING(f"  Cartella ignorata (formato non riconosciuto): {nome_cartella}")
                )
                continue

            zona_core, gruppo, reparto, nome_sq = parsed
            diario = _trova_diario(nome_sq, reparto, gruppo, zona_core, edizione)
            if not diario:
                err += 1
                if verbosity >= 2:
                    self.stdout.write(
                        self.style.ERROR(f"  Diario non trovato per: {nome_cartella}")
                    )
                continue

            # 3. Lista i file nella sottocartella
            files = self._lista_file(service, cartella["id"])
            importati = 0
            for f in files:
                filename = f["name"]
                modulo = modulo_map.get(filename)
                if not modulo:
                    # Cerca per corrispondenza parziale (Jotform aggiunge prefissi)
                    for k, v in modulo_map.items():
                        if filename in k or k in filename:
                            modulo = v
                            break
                if not modulo:
                    modulo = "impresa_1"  # fallback

                if dry_run:
                    importati += 1
                    continue

                _, created = Allegato.objects.get_or_create(
                    diario=diario,
                    drive_file_id=f["id"],
                    defaults={
                        "modulo": modulo,
                        "nome": filename,
                        "mime": f.get("mimeType", ""),
                        "dimensione": int(f.get("size", 0) or 0),
                        "stato_sync": StatoSync.CARICATO,
                        "caricato_da": None,
                    },
                )
                if created:
                    importati += 1

            ok += 1
            if verbosity >= 2:
                dr = " (dry-run)" if dry_run else ""
                self.stdout.write(f"  {nome_cartella}: {importati} file{dr}")

        self.stdout.write(
            self.style.SUCCESS(
                f"  Foto: {ok} cartelle elaborate, {err} non trovate."
            )
        )

    def _build_filename_modulo_map(self, wb) -> dict[str, str]:
        """Costruisce {filename → modulo} dagli URL foto dell'Excel."""
        ws = wb["Risposte EG"]
        mapping: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            for col_idx, modulo in [(29, "impresa_1"), (41, "impresa_2"), (46, "missione")]:
                cell_val = row[col_idx - 1]
                if not cell_val:
                    continue
                for url in str(cell_val).split("\n"):
                    url = url.strip()
                    if url.startswith("http"):
                        filename = url.rstrip("/").split("/")[-1]
                        if filename:
                            mapping[filename] = modulo
        return mapping

    @staticmethod
    def _lista_sottocartelle(service, folder_id: str) -> list[dict]:
        """Lista ricorsiva delle sottocartelle (una pagina alla volta)."""
        risultati = []
        page_token = None
        while True:
            kwargs: dict = {
                "q": (
                    f"'{folder_id}' in parents "
                    "and mimeType='application/vnd.google-apps.folder' "
                    "and trashed=false"
                ),
                "pageSize": 500,
                "fields": "nextPageToken, files(id, name)",
            }
            if page_token:
                kwargs["pageToken"] = page_token
            resp = service.files().list(**kwargs).execute()
            risultati.extend(resp.get("files", []))
            page_token = resp.get("nextPageToken")
            if not page_token:
                break
        return risultati

    @staticmethod
    def _lista_file(service, folder_id: str) -> list[dict]:
        """Lista i file (non cartelle) in una cartella Drive."""
        risultati = []
        page_token = None
        while True:
            kwargs: dict = {
                "q": (
                    f"'{folder_id}' in parents "
                    "and mimeType!='application/vnd.google-apps.folder' "
                    "and trashed=false"
                ),
                "pageSize": 200,
                "fields": "nextPageToken, files(id, name, size, mimeType, webViewLink)",
            }
            if page_token:
                kwargs["pageToken"] = page_token
            resp = service.files().list(**kwargs).execute()
            risultati.extend(resp.get("files", []))
            page_token = resp.get("nextPageToken")
            if not page_token:
                break
        return risultati
