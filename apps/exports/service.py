# apps/exports/service.py
"""Generazione PDF (WeasyPrint) ed Excel (openpyxl). Vedi docs sez. 11."""
from __future__ import annotations

import io
from pathlib import Path

from django.conf import settings

# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _get_pdf_template_html() -> str:
    """Restituisce il template HTML attivo (DB o file di default)."""
    from apps.exports.models import PdfTemplate

    try:
        tpl = PdfTemplate.objects.get(chiave="diario", attivo=True)
        return tpl.contenuto_html
    except PdfTemplate.DoesNotExist:
        path = Path(settings.BASE_DIR) / "templates" / "exports" / "diario.html"
        return path.read_text(encoding="utf-8")


def _compress_image_for_pdf(raw: bytes, mime: str, max_px: int = 480) -> tuple[bytes, str]:
    """Ridimensiona e comprime l'immagine per l'embedding nel PDF.

    Le immagini nel template sono 120×90pt (~250×188px a 150dpi).
    480px sul lato maggiore è il doppio della risoluzione di stampa — ottima qualità,
    dimensione file ridotta dell'80-90% rispetto agli originali.
    """
    import io

    from PIL import Image, UnidentifiedImageError

    try:
        img = Image.open(io.BytesIO(raw))
        img.load()
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        w, h = img.size
        if max(w, h) > max_px:
            ratio = max_px / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=82, optimize=True)
        return buf.getvalue(), "image/jpeg"
    except (UnidentifiedImageError, Exception):
        return raw, mime  # formato non supportato: usa l'originale


def _fetch_foto_drive(allegati, max_count: int = 6, budget_sec: float = 60.0) -> list[dict]:
    """Scarica le immagini da Drive, le comprime per il PDF e restituisce {nome, src} base64.

    Limita il numero di immagini (max_count) e il tempo totale (budget_sec).
    """
    import base64
    import io
    import time

    foto = []
    deadline = time.monotonic() + budget_sec

    for a in allegati[:max_count]:
        if time.monotonic() >= deadline:
            break
        if not a.drive_file_id or not a.mime.startswith("image/"):
            continue
        try:
            from googleapiclient.http import MediaIoBaseDownload

            from apps.storage_drive.service import _build_drive_service, _get_credenziali

            cred = _get_credenziali(a.diario.edizione)
            service = _build_drive_service(cred)
            req = service.files().get_media(fileId=a.drive_file_id)
            buf = io.BytesIO()
            dl = MediaIoBaseDownload(buf, req, chunksize=2 * 1024 * 1024)
            done = False
            while not done:
                if time.monotonic() >= deadline:
                    break
                _, done = dl.next_chunk()
            else:
                raw = buf.getvalue()
                compressed, out_mime = _compress_image_for_pdf(raw, a.mime)
                b64 = base64.b64encode(compressed).decode()
                foto.append({"nome": a.nome, "src": f"data:{out_mime};base64,{b64}"})
        except Exception:
            pass  # immagine non disponibile: la saltiamo nel PDF
    return foto


def genera_pdf_diario(diario, include_relazione: bool = False) -> bytes:
    """Genera il PDF del diario.

    include_relazione: se True include la Relazione finale CRP (modulo 6).
    Non passare True per i Capi Squadriglia — docs §15.
    """
    import weasyprint

    from apps.diaries.models import Allegato
    from apps.siteconfig.models import Impostazioni

    imp = Impostazioni.get()

    imprese = list(diario.imprese.prefetch_related("posti_azione", "esiti_specialita"))
    all_foto = list(Allegato.objects.filter(
        diario=diario, mime__startswith="image/"
    ).only("pk", "diario_id", "modulo", "drive_file_id", "mime", "nome"))

    for imp_obj in imprese:
        modulo_key = f"impresa_{imp_obj.numero}"
        imp_obj.pdf_foto = _fetch_foto_drive(
            [a for a in all_foto if a.modulo == modulo_key]
        )

    missione = getattr(diario, "missione", None)
    missione_foto = _fetch_foto_drive(
        [a for a in all_foto if a.modulo == "missione"]
    ) if missione else []

    relazione_finale = None
    if include_relazione:
        relazione_finale = getattr(diario, "relazione_finale", None)

    context = {
        "diario": diario,
        "anagrafica": getattr(diario, "anagrafica", None),
        "presentazione": getattr(diario, "presentazione", None),
        "imprese": imprese,
        "missione": missione,
        "missione_foto": missione_foto,
        "relazione_finale": relazione_finale,
        "titolo_piattaforma": imp.titolo,
        "sottotitolo_piattaforma": imp.sottotitolo,
    }

    template_html = _get_pdf_template_html()

    from django.template import Context, Template
    rendered = Template(template_html).render(Context(context))

    return weasyprint.HTML(string=rendered, base_url=str(settings.BASE_DIR)).write_pdf()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def genera_excel_edizione(edizione) -> bytes:
    """Genera un Excel con una riga per diario e fogli per zona. Vedi docs sez. 11."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws_riepilogo = wb.active
    ws_riepilogo.title = "Riepilogo"

    header = [
        "Zona", "Gruppo", "Reparto", "Squadriglia", "Tipo",
        "CSQ Cognome", "CSQ Nome", "CRP Cognome", "CRP Nome",
        "Stato", "Scadenza", "Specialità",
        "Esito", "Pubblicato", "Note valutazione",
    ]
    ws_riepilogo.append(header)
    # Stile intestazione
    fill = PatternFill("solid", fgColor="5AA02C")
    for i, _ in enumerate(header, 1):
        cell = ws_riepilogo.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    diari = (
        edizione.diari
        .select_related("squadriglia__reparto__gruppo__zona", "csq", "crp")
        .prefetch_related("anagrafica", "valutazione")
        .order_by(
            "squadriglia__reparto__gruppo__zona__nome",
            "squadriglia__reparto__gruppo__nome",
            "squadriglia__nome",
        )
    )

    zone_viste: dict[str, object] = {}

    for d in diari:
        sq = d.squadriglia
        zona_nome = sq.reparto.gruppo.zona.nome
        ana = getattr(d, "anagrafica", None)
        val = getattr(d, "valutazione", None)

        riga = [
            zona_nome,
            sq.reparto.gruppo.nome,
            sq.reparto.nome,
            sq.nome,
            d.get_tipo_display(),
            d.csq.cognome if d.csq else "",
            d.csq.nome if d.csq else "",
            d.crp.cognome if d.crp else "",
            d.crp.nome if d.crp else "",
            d.get_stato_display(),
            str(d.scadenza_effettiva() or ""),
            ana.specialita if ana else "",
            val.get_esito_display() if val and val.esito else "",
            "Sì" if d.pubblicato else "No",
            val.note if val else "",
        ]
        ws_riepilogo.append(riga)

        # Un foglio per zona
        if zona_nome not in zone_viste:
            ws_zona = wb.create_sheet(zona_nome[:31])
            ws_zona.append(header)
            zone_viste[zona_nome] = ws_zona
        zone_viste[zona_nome].append(riga)

    # Larghezze colonne automatiche
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Export riassuntivo diari (xlsx / ods / csv)
# ---------------------------------------------------------------------------

def _header_export_diari() -> list[str]:
    return [
        # Identificazione
        "Zona", "Gruppo", "Reparto", "Squadriglia", "Tipo", "Stato", "Scadenza",
        # Anagrafica CRP
        "CRP Cognome", "CRP Nome", "CRP Email", "CRP Cellulare",
        # Anagrafica CSQ
        "CSQ Cognome", "CSQ Nome", "CSQ Email", "CSQ Cellulare",
        # Specialità
        "Specialità", "Partecipa all'evento",
        # Presentazione
        "Cosa sappiamo fare", "Membri",
        # Impresa 1
        "Impresa 1 — Titolo", "Impresa 1 — Data inizio", "Impresa 1 — Data fine",
        "Impresa 1 — Perché", "Impresa 1 — Come", "Impresa 1 — Cosa",
        "Impresa 1 — Posti d'azione", "Impresa 1 — Esiti specialità",
        # Impresa 2
        "Impresa 2 — Titolo", "Impresa 2 — Data inizio", "Impresa 2 — Data fine",
        "Impresa 2 — Perché", "Impresa 2 — Come", "Impresa 2 — Cosa",
        "Impresa 2 — Posti d'azione", "Impresa 2 — Esiti specialità",
        # Missione
        "Missione — Titolo", "Missione — Data", "Missione — Descrizione",
        "Missione — Posti d'azione",
        # Relazione finale (CRP)
        "RF — Sintesi 1ª impresa", "RF — Sintesi 2ª impresa",
        "RF — Sintesi missione", "RF — Considerazioni", "RF — Specialità conquistata",
        # Valutazione (solo staff — inclusa condizionalmente)
        "Valutazione — Esito", "Valutazione — Note",
        # Drive
        "Link cartella Drive",
    ]


def _riga_export_diario(diario, mostra_valutazione: bool) -> list:
    sq = diario.squadriglia
    ana = getattr(diario, "anagrafica", None)
    pres = getattr(diario, "presentazione", None)
    missione = getattr(diario, "missione", None)
    rf = getattr(diario, "relazione_finale", None)
    val = getattr(diario, "valutazione", None)

    def _d(v) -> str:
        if v is None:
            return ""
        return v.strftime("%d/%m/%Y") if hasattr(v, "strftime") else str(v)

    def _posti_azione(qs) -> str:
        parts = []
        for p in qs.all():
            chi = getattr(p, "chi", "")
            cosa = getattr(p, "cosa", "") or getattr(p, "descrizione", "")
            if chi and cosa:
                parts.append(f"{chi}: {cosa}")
            elif cosa:
                parts.append(cosa)
            elif chi:
                parts.append(chi)
        return " | ".join(parts)

    def _esiti(qs) -> str:
        parts = []
        for e in qs.all():
            chi_str = f" ({e.chi})" if getattr(e, "chi", "") else ""
            parts.append(f"{e.get_tipo_display()} — {e.nome}{chi_str} [{e.get_stato_display()}]")
        return " | ".join(parts)

    def _impresa_cols(numero: int) -> list:
        try:
            imp = next(i for i in diario._imprese_cache if i.numero == numero)
        except (StopIteration, AttributeError):
            return [""] * 8
        pa = _posti_azione(imp.posti_azione)
        es = _esiti(imp.esiti_specialita)
        return [
            imp.titolo, _d(imp.data_inizio), _d(imp.data_fine),
            imp.perche, imp.come, imp.cosa, pa, es,
        ]

    def _membri() -> str:
        if pres is None:
            return ""
        parts = []
        for m in pres.membri.all():
            ruolo = m.get_ruolo_display() if m.ruolo else ""
            sent = m.get_sentiero_display() if m.sentiero else ""
            dettaglio = ", ".join(filter(None, [ruolo, sent]))
            parts.append(f"{m.nome} ({dettaglio})" if dettaglio else m.nome)
        return " | ".join(parts)

    drive_url = ""
    if diario.drive_folder_allegati_id:
        drive_url = f"https://drive.google.com/drive/folders/{diario.drive_folder_allegati_id}"

    riga: list = [
        # Identificazione
        sq.reparto.gruppo.zona.nome,
        sq.reparto.gruppo.nome,
        sq.reparto.nome,
        sq.nome,
        diario.get_tipo_display(),
        diario.get_stato_display(),
        _d(diario.scadenza_effettiva()) if hasattr(diario, "scadenza_effettiva") else "",
        # Anagrafica CRP
        (ana.crp_cognome if ana else ""),
        (ana.crp_nome if ana else ""),
        (ana.crp_email if ana else ""),
        (ana.crp_cell if ana else ""),
        # Anagrafica CSQ
        (ana.csq_cognome if ana else ""),
        (ana.csq_nome if ana else ""),
        (ana.csq_email if ana else ""),
        (ana.csq_cell if ana else ""),
        # Specialità
        (ana.specialita if ana else ""),
        ("Sì" if ana and ana.partecipa_evento else "No") if ana else "",
        # Presentazione
        (pres.cosa_sappiamo_fare if pres else ""),
        _membri(),
    ]
    riga += _impresa_cols(1)
    riga += _impresa_cols(2)
    riga += [
        # Missione
        (missione.titolo if missione else ""),
        _d(missione.data) if missione else "",
        (missione.descrizione_svolgimento if missione else ""),
        _posti_azione(missione.posti_azione_missione) if missione else "",
        # Relazione finale
        (rf.sintesi_impresa_1 if rf else ""),
        (rf.sintesi_impresa_2 if rf else ""),
        (rf.sintesi_missione if rf else ""),
        (rf.considerazioni if rf else ""),
        ("Sì" if rf and rf.specialita_conquistata else "") if rf else "",
        # Valutazione
        (val.get_esito_display() if val and val.esito else "") if mostra_valutazione else "",
        (val.note if val else "") if mostra_valutazione else "",
        # Drive
        drive_url,
    ]
    return riga


def costruisci_tabella_diari(qs, user) -> tuple[list[str], list[list]]:
    """Costruisce header + righe per l'export riassuntivo dei diari.

    mostra_valutazione: True per staff_plancia e superuser.
    """
    mostra_valutazione = user.is_staff_plancia or user.is_superuser

    diari = list(
        qs.select_related(
            "squadriglia__reparto__gruppo__zona",
            "csq", "crp",
            "anagrafica", "presentazione",
            "missione",
            "relazione_finale",
            "valutazione",
        ).prefetch_related(
            "presentazione__membri",
            "imprese__posti_azione",
            "imprese__esiti_specialita",
            "missione__posti_azione_missione",
        )
    )

    for d in diari:
        d._imprese_cache = list(d.imprese.all())

    header = _header_export_diari()
    righe = [_riga_export_diario(d, mostra_valutazione) for d in diari]
    return header, righe


def genera_export_diari(qs, user, formato: str) -> tuple[bytes, str, str]:
    """Genera l'export riassuntivo dei diari nel formato richiesto.

    Restituisce (contenuto, content_type, filename).
    formato: 'xlsx' | 'ods' | 'csv'
    """
    header, righe = costruisci_tabella_diari(qs, user)

    if formato == "xlsx":
        return _write_xlsx(header, righe), (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), "export_diari.xlsx"
    if formato == "ods":
        return _write_ods(header, righe), (
            "application/vnd.oasis.opendocument.spreadsheet"
        ), "export_diari.ods"
    # csv (default)
    return _write_csv(header, righe), "text/csv; charset=utf-8", "export_diari.csv"


def _write_xlsx(header: list[str], righe: list[list]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diari"
    ws.append(header)
    fill = PatternFill("solid", fgColor="5AA02C")
    for i, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for riga in righe:
        ws.append([str(v) if v is not None else "" for v in riga])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_csv(header: list[str], righe: list[list]) -> bytes:
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for riga in righe:
        writer.writerow([str(v) if v is not None else "" for v in riga])
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _write_ods(header: list[str], righe: list[list]) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import Style, TableCellProperties, TextProperties
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = OpenDocumentSpreadsheet()

    # Stile intestazione verde
    style_header = Style(name="HeaderCell", family="table-cell")
    style_header.addElement(TextProperties(fontweight="bold", color="#ffffff"))
    style_header.addElement(TableCellProperties(backgroundcolor="#5AA02C"))
    doc.automaticstyles.addElement(style_header)

    table = Table(name="Diari")
    doc.spreadsheet.addElement(table)

    def _cell(val: str, style_name: str | None = None) -> TableCell:
        attrs: dict = {"valuetype": "string"}
        if style_name:
            attrs["stylename"] = style_name
        tc = TableCell(**attrs)
        tc.addElement(P(text=val))
        return tc

    tr = TableRow()
    for h in header:
        tr.addElement(_cell(h, "HeaderCell"))
    table.addElement(tr)

    for riga in righe:
        tr = TableRow()
        for v in riga:
            tr.addElement(_cell(str(v) if v is not None else ""))
        table.addElement(tr)

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()
